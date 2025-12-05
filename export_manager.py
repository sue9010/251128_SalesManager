import os
import time
import shutil
import openpyxl
from config import Config

class ExportManager:
    def __init__(self):
        pass

    def export_quote_to_pdf(self, client_info, quote_info, items):
        """견적서 PDF 변환 (기존 코드 유지)"""
        try:
            import win32com.client
            import pythoncom
        except ImportError:
            return False, "PDF 변환을 위해 pywin32 라이브러리가 필요합니다.\n설치: pip install pywin32"

        temp_xlsx_path = None
        excel = None
        wb_opened = None

        try:
            client_name = quote_info['client_name']
            mgmt_no = quote_info['mgmt_no']
            
            country = str(client_info.get("국가", "Korea")).upper()
            is_kr = any(x in country for x in ["대한민국", "KR", "한국", "KOREA"])
            
            template_filename = "Quotation_KR.xlsx" if is_kr else "Quotation_EN.xlsx"
            template_dir = os.path.join(Config.DEFAULT_ATTACHMENT_ROOT, "forms")
            template_path = os.path.join(template_dir, template_filename)

            if not os.path.exists(template_path):
                return False, f"견적서 양식 파일을 찾을 수 없습니다.\n경로: {template_path}"

            wb = openpyxl.load_workbook(template_path)
            ws = wb.active

            ws["B7"] = client_name
            ws["B8"] = str(client_info.get("담당자", ""))
            ws["B9"] = str(client_info.get("전화번호", ""))
            ws["B10"] = str(client_info.get("주소", ""))

            ws["G10"] = quote_info['date']
            ws["G11"] = mgmt_no

            start_row = 14
            for i, item in enumerate(items):
                if i >= 10: break
                r_idx = start_row + i
                
                ws[f"B{r_idx}"] = item.get('item', '')
                ws[f"C{r_idx}"] = item.get('model', '')
                ws[f"D{r_idx}"] = item.get('desc', '')
                ws[f"E{r_idx}"] = item.get('qty', 0)
                ws[f"F{r_idx}"] = item.get('price', 0)
                ws[f"G{r_idx}"] = item.get('amount', 0)

            ws["B28"] = quote_info.get('note', '')

            desktop = os.path.join(os.path.expanduser("~"), "Desktop")
            
            if is_kr:
                pdf_filename = f"견적서_{client_name}_{mgmt_no}.pdf"
            else:
                pdf_filename = f"Quotation_{client_name}_{mgmt_no}.pdf"
            
            temp_xlsx_name = f"temp_{mgmt_no}.xlsx"
            temp_xlsx_path = os.path.join(desktop, temp_xlsx_name)
            pdf_path = os.path.join(desktop, pdf_filename)
            
            wb.save(temp_xlsx_path)
            wb.close()

            pythoncom.CoInitialize()
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False 
            excel.ScreenUpdating = False

            abs_xlsx_path = os.path.abspath(temp_xlsx_path)
            wb_opened = excel.Workbooks.Open(abs_xlsx_path)
            
            try:
                wb_opened.ExportAsFixedFormat(0, pdf_path)
            except Exception as e:
                return False, f"PDF 변환 실패 (엑셀 오류): {e}"
            
            wb_opened.Close(SaveChanges=False)
            wb_opened = None
            
            return True, pdf_path

        except Exception as e:
            return False, f"오류 발생: {str(e)}"
            
        finally:
            if wb_opened:
                try: wb_opened.Close(False)
                except: pass
            if excel:
                try: excel.Quit(); del excel
                except: pass
            if temp_xlsx_path and os.path.exists(temp_xlsx_path):
                try: os.remove(temp_xlsx_path)
                except: pass

    def export_order_request_to_pdf(self, client_info, order_info, items):
        """출고요청서 PDF 변환 (기존 코드 유지)"""
        try:
            import win32com.client
            import pythoncom
        except ImportError:
            return False, "PDF 변환을 위해 pywin32 라이브러리가 필요합니다."

        temp_xlsx_path = None
        excel = None
        wb_opened = None

        try:
            template_path = Config.ORDER_REQUEST_FORM_PATH
            if not os.path.exists(template_path):
                return False, f"출고요청서 양식 파일을 찾을 수 없습니다.\n경로: {template_path}"

            wb = openpyxl.load_workbook(template_path, keep_vba=True)
            ws = wb.active

            ws["C1"] = order_info.get('type', '')
            ws["C2"] = order_info.get('mgmt_no', '')
            ws["C3"] = order_info.get('date', '')
            
            ws["C7"] = str(client_info.get("주소", ""))
            
            ws["E1"] = order_info.get('client_name', '')
            ws["E3"] = str(client_info.get("국가", ""))
            ws["E4"] = str(client_info.get("전화번호", ""))
            ws["E5"] = str(client_info.get("운송계정", ""))
            ws["E6"] = str(client_info.get("운송방법", ""))

            start_row = 10
            for i, item in enumerate(items):
                if i >= 10: break 
                r_idx = start_row + i
                
                ws[f"C{r_idx}"] = item.get('item', '')
                ws[f"D{r_idx}"] = item.get('model', '')
                ws[f"E{r_idx}"] = item.get('desc', '')
                ws[f"F{r_idx}"] = item.get('qty', 0)

            ws["B21"] = order_info.get('req_note', '')
            ws["B24"] = str(client_info.get("특이사항", ""))

            target_dir = os.path.join(Config.DEFAULT_ATTACHMENT_ROOT, "출고요청서")
            if not os.path.exists(target_dir):
                try: os.makedirs(target_dir)
                except: pass

            client_safe = "".join([c for c in order_info['client_name'] if c.isalnum() or c in (' ', '_')]).strip()
            pdf_filename = f"출고요청서_{client_safe}_{order_info['mgmt_no']}.pdf"
            
            desktop = os.path.join(os.path.expanduser("~"), "Desktop")
            temp_xlsx_name = f"temp_req_{order_info['mgmt_no']}.xlsm"
            temp_xlsx_path = os.path.join(desktop, temp_xlsx_name)
            
            if os.path.exists(target_dir):
                pdf_path = os.path.join(target_dir, pdf_filename)
            else:
                pdf_path = os.path.join(desktop, pdf_filename)

            wb.save(temp_xlsx_path)
            wb.close()

            pythoncom.CoInitialize()
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            excel.ScreenUpdating = False

            abs_xlsx_path = os.path.abspath(temp_xlsx_path)
            wb_opened = excel.Workbooks.Open(abs_xlsx_path)
            
            try:
                wb_opened.ExportAsFixedFormat(0, pdf_path)
            except Exception as e:
                return False, f"PDF 변환 실패: {e}"
            
            wb_opened.Close(SaveChanges=False)
            wb_opened = None
            
            return True, pdf_path

        except Exception as e:
            return False, f"오류 발생: {str(e)}"
            
        finally:
            if wb_opened:
                try: wb_opened.Close(False)
                except: pass
            if excel:
                try: excel.Quit(); del excel
                except: pass
            if temp_xlsx_path and os.path.exists(temp_xlsx_path):
                try: os.remove(temp_xlsx_path)
                except: pass

    def export_pi_to_pdf(self, client_info, order_info, items):
        """PI 발행 (기존 코드 유지)"""
        try:
            import win32com.client
            import pythoncom
        except ImportError:
            return False, "PDF 변환을 위해 pywin32 라이브러리가 필요합니다."

        temp_xlsx_path = None
        excel = None
        wb_opened = None

        try:
            template_path = os.path.join(Config.DEFAULT_ATTACHMENT_ROOT, "forms", "PI.xlsx")
            
            if not os.path.exists(template_path):
                return False, f"PI 양식 파일을 찾을 수 없습니다.\n경로: {template_path}"

            wb = openpyxl.load_workbook(template_path)
            ws = wb.active

            ws["B5"] = order_info.get('client_name', '')
            ws["B6"] = str(client_info.get("담당자", ""))
            ws["B7"] = str(client_info.get("전화번호", ""))
            ws["B8"] = str(client_info.get("주소", ""))
            ws["G9"] = str(client_info.get("국가", ""))

            ws["G5"] = order_info.get('date', '')
            ws["G6"] = order_info.get('mgmt_no', '')
            ws["G7"] = order_info.get('po_no', '')

            start_row = 12
            for i, item in enumerate(items):
                if i >= 10: break
                r_idx = start_row + i
                
                ws[f"B{r_idx}"] = item.get('item', '')
                ws[f"C{r_idx}"] = item.get('model', '')
                ws[f"D{r_idx}"] = item.get('desc', '')
                ws[f"E{r_idx}"] = item.get('qty', 0)
                ws[f"F{r_idx}"] = item.get('price', 0)
                ws[f"G{r_idx}"] = item.get('amount', 0)

            desktop = os.path.join(os.path.expanduser("~"), "Desktop")
            server_pi_folder = os.path.join(Config.DEFAULT_ATTACHMENT_ROOT, "Proforma Invoice")
            
            if not os.path.exists(server_pi_folder):
                try: os.makedirs(server_pi_folder)
                except: pass

            client_safe = "".join([c for c in order_info['client_name'] if c.isalnum() or c in (' ', '_', '-')]).strip()
            pdf_filename = f"PI_{client_safe}_{order_info['mgmt_no']}.pdf"
            
            temp_xlsx_name = f"temp_PI_{order_info['mgmt_no']}.xlsx"
            temp_xlsx_path = os.path.join(desktop, temp_xlsx_name)
            
            desktop_pdf_path = os.path.join(desktop, pdf_filename)
            server_pdf_path = os.path.join(server_pi_folder, pdf_filename)

            wb.save(temp_xlsx_path)
            wb.close()

            pythoncom.CoInitialize()
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            excel.ScreenUpdating = False

            abs_xlsx_path = os.path.abspath(temp_xlsx_path)
            wb_opened = excel.Workbooks.Open(abs_xlsx_path)
            
            try:
                wb_opened.ExportAsFixedFormat(0, desktop_pdf_path)
            except Exception as e:
                return False, f"PDF 변환 실패: {e}"
            
            wb_opened.Close(SaveChanges=False)
            wb_opened = None
            
            copied_to_server = False
            if os.path.exists(server_pi_folder):
                try:
                    shutil.copy2(desktop_pdf_path, server_pdf_path)
                    copied_to_server = True
                except Exception as e:
                    print(f"서버 복사 실패: {e}")

            msg = f"바탕화면에 저장되었습니다.\n{desktop_pdf_path}"
            if copied_to_server:
                msg += f"\n\n서버에도 저장되었습니다.\n{server_pdf_path}"

            return True, msg

        except Exception as e:
            return False, f"오류 발생: {str(e)}"
            
        finally:
            if wb_opened:
                try: wb_opened.Close(False)
                except: pass
            if excel:
                try: excel.Quit(); del excel
                except: pass
            if temp_xlsx_path and os.path.exists(temp_xlsx_path):
                try: os.remove(temp_xlsx_path)
                except: pass

    def export_ci_to_pdf(self, client_info, order_info, items):
        """
        CI.xlsx 양식에 데이터를 채우고 바탕화면과 서버 폴더에 PDF로 저장합니다.
        """
        try:
            import win32com.client
            import pythoncom
        except ImportError:
            return False, "PDF 변환을 위해 pywin32 라이브러리가 필요합니다."

        temp_xlsx_path = None
        excel = None
        wb_opened = None

        try:
            template_path = os.path.join(Config.DEFAULT_ATTACHMENT_ROOT, "forms", "CI.xlsx")
            
            if not os.path.exists(template_path):
                return False, f"CI 양식 파일을 찾을 수 없습니다.\n경로: {template_path}"

            wb = openpyxl.load_workbook(template_path)
            ws = wb.active

            # 고객 정보
            ws["A9"] = order_info.get('client_name', '')
            ws["A10"] = str(client_info.get("주소", ""))
            ws["A12"] = str(client_info.get("전화번호", ""))
            ws["A13"] = str(client_info.get("담당자", ""))
            ws["E13"] = str(client_info.get("국가", ""))

            # 주문 정보
            ws["E9"] = order_info.get('mgmt_no', '')
            ws["E11"] = order_info.get('date', '')

            # [수정] 발주서 번호 일괄 처리
            # items에서 모든 발주서 번호를 수집하여 쉼표로 연결
            po_list = []
            
            # 우선 order_info에 단일 po_no가 있다면 추가
            if order_info.get('po_no'):
                po_list.append(str(order_info.get('po_no')))
            
            # items 내의 po_no들도 확인
            for item in items:
                p = str(item.get('po_no', '')).strip()
                if p and p != "nan" and p not in po_list:
                    po_list.append(p)
            
            # 중복 제거 및 연결
            final_po_str = ", ".join(list(dict.fromkeys(po_list))) # 순서 유지하며 중복 제거
            ws["E10"] = final_po_str

            # 품목 정보 & 시리얼 번호 수집
            serial_list = []
            
            start_row = 16
            for i, item in enumerate(items):
                if i >= 10: break
                r_idx = start_row + i
                
                ws[f"A{r_idx}"] = item.get('model', '')
                ws[f"B{r_idx}"] = item.get('desc', '')
                ws[f"C{r_idx}"] = item.get('qty', 0)
                ws[f"D{r_idx}"] = item.get('currency', '')
                ws[f"E{r_idx}"] = item.get('price', 0)
                ws[f"H{r_idx}"] = item.get('currency', '')
                ws[f"I{r_idx}"] = item.get('amount', 0)
                
                # 시리얼 번호 수집
                sn = str(item.get('serial', '')).strip()
                if sn and sn != "-" and sn != "nan":
                    serial_list.append(sn)

            # [수정] A32에 시리얼 번호 목록 입력
            if serial_list:
                ws["A32"] = ", ".join(serial_list)

            # 경로 설정
            desktop = os.path.join(os.path.expanduser("~"), "Desktop")
            server_ci_folder = os.path.join(Config.DEFAULT_ATTACHMENT_ROOT, "Commercial Invoice")
            
            if not os.path.exists(server_ci_folder):
                try: os.makedirs(server_ci_folder)
                except: pass

            client_safe = "".join([c for c in order_info['client_name'] if c.isalnum() or c in (' ', '_', '-')]).strip()
            pdf_filename = f"CI_{client_safe}_{order_info['mgmt_no']}.pdf"
            
            temp_xlsx_name = f"temp_CI_{order_info['mgmt_no']}.xlsx"
            temp_xlsx_path = os.path.join(desktop, temp_xlsx_name)
            
            desktop_pdf_path = os.path.join(desktop, pdf_filename)
            server_pdf_path = os.path.join(server_ci_folder, pdf_filename)

            wb.save(temp_xlsx_path)
            wb.close()

            # PDF 변환
            pythoncom.CoInitialize()
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            excel.ScreenUpdating = False

            abs_xlsx_path = os.path.abspath(temp_xlsx_path)
            wb_opened = excel.Workbooks.Open(abs_xlsx_path)
            
            try:
                wb_opened.ExportAsFixedFormat(0, desktop_pdf_path)
            except Exception as e:
                return False, f"PDF 변환 실패: {e}"
            
            wb_opened.Close(SaveChanges=False)
            wb_opened = None
            
            copied_to_server = False
            if os.path.exists(server_ci_folder):
                try:
                    shutil.copy2(desktop_pdf_path, server_pdf_path)
                    copied_to_server = True
                except Exception as e:
                    print(f"서버 복사 실패: {e}")

            msg = f"바탕화면에 저장되었습니다.\n{desktop_pdf_path}"
            if copied_to_server:
                msg += f"\n\n서버에도 저장되었습니다.\n{server_pdf_path}"

            return True, msg

        except Exception as e:
            return False, f"오류 발생: {str(e)}"
            
        finally:
            if wb_opened:
                try: wb_opened.Close(False)
                except: pass
            if excel:
                try: excel.Quit(); del excel
                except: pass
            if temp_xlsx_path and os.path.exists(temp_xlsx_path):
                try: os.remove(temp_xlsx_path)
                except: pass