import getpass
import json
import os
import shutil
from datetime import datetime

import pandas as pd
import openpyxl

from config import Config


class DataManager:
    def __init__(self):
        self.df_clients = pd.DataFrame(columns=Config.CLIENT_COLUMNS)
        self.df_data = pd.DataFrame(columns=Config.DATA_COLUMNS)
        self.df_log = pd.DataFrame(columns=Config.LOG_COLUMNS)
        self.df_memo = pd.DataFrame(columns=Config.MEMO_COLUMNS)
        self.df_memo_log = pd.DataFrame(columns=Config.MEMO_LOG_COLUMNS)
        
        self.current_excel_path = Config.DEFAULT_EXCEL_PATH
        self.attachment_root = Config.DEFAULT_ATTACHMENT_ROOT
        self.production_request_path = Config.DEFAULT_PRODUCTION_REQUEST_PATH
        
        # [수정] 이 부분이 누락되었을 수 있습니다. 확실하게 초기화합니다.
        self.order_request_dir = Config.DEFAULT_ORDER_REQUEST_DIR 
        
        self.current_theme = "Dark"
        self.is_dev_mode = False
        
        self.last_file_timestamp = 0.0
        
        self.load_config()

    def load_config(self):
        if os.path.exists(Config.CONFIG_FILENAME):
            try:
                with open(Config.CONFIG_FILENAME, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    self.current_excel_path = data.get("excel_path", Config.DEFAULT_EXCEL_PATH)
                    self.current_theme = data.get("theme", "Dark")
                    self.attachment_root = data.get("attachment_root", Config.DEFAULT_ATTACHMENT_ROOT)
                    self.production_request_path = data.get("production_request_path", Config.DEFAULT_PRODUCTION_REQUEST_PATH)
                    # 설정 파일에서 로드
                    self.order_request_dir = data.get("order_request_dir", Config.DEFAULT_ORDER_REQUEST_DIR)
            except: pass

    def save_config(self, new_path=None, new_theme=None, new_attachment_dir=None, new_prod_path=None, new_order_req_dir=None):
        if new_path: self.current_excel_path = new_path
        if new_theme: self.current_theme = new_theme
        if new_attachment_dir: self.attachment_root = new_attachment_dir
        if new_prod_path: self.production_request_path = new_prod_path
        if new_order_req_dir: self.order_request_dir = new_order_req_dir
        
        data = {
            "excel_path": self.current_excel_path,
            "theme": self.current_theme,
            "attachment_root": self.attachment_root,
            "production_request_path": self.production_request_path,
            "order_request_dir": self.order_request_dir
        }
        try:
            with open(Config.CONFIG_FILENAME, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=4)
        except Exception as e:
            print(f"설정 저장 실패: {e}")

    # ... (나머지 메서드들은 기존과 동일하게 유지) ...
    # ... (load_data, check_for_external_changes, _execute_transaction, _create_log_entry 등) ...

    def load_data(self):
        if not os.path.exists(self.current_excel_path):
            return False, "파일이 존재하지 않습니다."

        try:
            current_mtime = os.path.getmtime(self.current_excel_path)
            
            with pd.ExcelFile(self.current_excel_path) as xls:
                if Config.SHEET_CLIENTS in xls.sheet_names:
                    self.df_clients = pd.read_excel(xls, Config.SHEET_CLIENTS)
                    self.df_clients.columns = self.df_clients.columns.str.strip()
                
                if Config.SHEET_DATA in xls.sheet_names:
                    self.df_data = pd.read_excel(xls, Config.SHEET_DATA)
                    self.df_data.columns = self.df_data.columns.str.strip()
                
                if Config.SHEET_LOG in xls.sheet_names:
                    self.df_log = pd.read_excel(xls, Config.SHEET_LOG)
                else:
                    self.df_log = pd.DataFrame(columns=Config.LOG_COLUMNS)

                if Config.SHEET_MEMO in xls.sheet_names:
                    self.df_memo = pd.read_excel(xls, Config.SHEET_MEMO)
                else:
                    self.df_memo = pd.DataFrame(columns=Config.MEMO_COLUMNS)

                if Config.SHEET_MEMO_LOG in xls.sheet_names:
                    self.df_memo_log = pd.read_excel(xls, Config.SHEET_MEMO_LOG)
                else:
                    self.df_memo_log = pd.DataFrame(columns=Config.MEMO_LOG_COLUMNS)

            self._preprocess_data()
            self.last_file_timestamp = current_mtime
            
            return True, "데이터 로드 완료"
        except Exception as e:
            return False, f"오류 발생: {e}"

    def check_for_external_changes(self):
        if not os.path.exists(self.current_excel_path): return False
        try:
            current_mtime = os.path.getmtime(self.current_excel_path)
            if current_mtime > self.last_file_timestamp: return True
        except OSError: pass
        return False

    def _execute_transaction(self, update_logic_func):
        if not os.path.exists(self.current_excel_path):
            return False, "엑셀 파일이 존재하지 않습니다."

        try:
            with pd.ExcelFile(self.current_excel_path) as xls:
                if Config.SHEET_CLIENTS in xls.sheet_names:
                    temp_clients = pd.read_excel(xls, Config.SHEET_CLIENTS)
                    temp_clients.columns = temp_clients.columns.str.strip()
                else: temp_clients = pd.DataFrame(columns=Config.CLIENT_COLUMNS)

                if Config.SHEET_DATA in xls.sheet_names:
                    temp_data = pd.read_excel(xls, Config.SHEET_DATA)
                    temp_data.columns = temp_data.columns.str.strip()
                else: temp_data = pd.DataFrame(columns=Config.DATA_COLUMNS)

                if Config.SHEET_LOG in xls.sheet_names:
                    temp_log = pd.read_excel(xls, Config.SHEET_LOG)
                else: temp_log = pd.DataFrame(columns=Config.LOG_COLUMNS)

                if Config.SHEET_MEMO in xls.sheet_names:
                    temp_memo = pd.read_excel(xls, Config.SHEET_MEMO)
                else: temp_memo = pd.DataFrame(columns=Config.MEMO_COLUMNS)
                
                if Config.SHEET_MEMO_LOG in xls.sheet_names:
                    temp_memo_log = pd.read_excel(xls, Config.SHEET_MEMO_LOG)
                else: temp_memo_log = pd.DataFrame(columns=Config.MEMO_LOG_COLUMNS)

            for col in Config.DATA_COLUMNS:
                if col not in temp_data.columns: temp_data[col] = "-"
            temp_data = temp_data.fillna("-")
            temp_clients = temp_clients.fillna("-")

            dfs = {
                "clients": temp_clients, "data": temp_data, "log": temp_log,
                "memo": temp_memo, "memo_log": temp_memo_log
            }
            
            success, msg = update_logic_func(dfs)
            if not success: return False, msg

            with pd.ExcelWriter(self.current_excel_path, engine="openpyxl") as writer:
                dfs["clients"].to_excel(writer, sheet_name=Config.SHEET_CLIENTS, index=False)
                dfs["data"].to_excel(writer, sheet_name=Config.SHEET_DATA, index=False)
                dfs["log"].to_excel(writer, sheet_name=Config.SHEET_LOG, index=False)
                dfs["memo"].to_excel(writer, sheet_name=Config.SHEET_MEMO, index=False)
                dfs["memo_log"].to_excel(writer, sheet_name=Config.SHEET_MEMO_LOG, index=False)

            self.load_data()
            return True, "저장되었습니다."

        except PermissionError:
            return False, "엑셀 파일이 열려있습니다. 파일을 닫고 다시 시도해주세요."
        except Exception as e:
            return False, f"트랜잭션 오류: {e}"

    def _create_log_entry(self, action, details):
        try: user = getpass.getuser()
        except: user = "Unknown"
        return {
            "일시": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "작업자": user,
            "구분": action,
            "상세내용": details
        }

    def _preprocess_data(self):
        for col in Config.DATA_COLUMNS:
            if col not in self.df_data.columns: self.df_data[col] = "-"
        
        self.df_data = self.df_data.fillna("-")
        self.df_clients = self.df_clients.fillna("-")
        
        num_cols = ["수량", "단가", "환율", "세율(%)", "공급가액", "세액", "합계금액", "기수금액", "미수금액"]
        for col in num_cols:
            if col in self.df_data.columns:
                self.df_data[col] = pd.to_numeric(self.df_data[col], errors='coerce').fillna(0)

        date_cols = ["견적일", "수주일", "출고예정일", "출고일", "선적일", "입금완료일", "세금계산서발행일"]
        for col in date_cols:
            if col in self.df_data.columns:
                self.df_data[col] = pd.to_datetime(self.df_data[col], errors='coerce', format='mixed').dt.strftime("%Y-%m-%d")
                self.df_data[col] = self.df_data[col].fillna("-")

    def save_to_excel(self):
        try:
            with pd.ExcelWriter(self.current_excel_path, engine="openpyxl") as writer:
                self.df_clients.to_excel(writer, sheet_name=Config.SHEET_CLIENTS, index=False)
                self.df_data.to_excel(writer, sheet_name=Config.SHEET_DATA, index=False)
                self.df_log.to_excel(writer, sheet_name=Config.SHEET_LOG, index=False)
                self.df_memo.to_excel(writer, sheet_name=Config.SHEET_MEMO, index=False)
                self.df_memo_log.to_excel(writer, sheet_name=Config.SHEET_MEMO_LOG, index=False)
            return True, "저장 완료"
        except PermissionError:
            return False, "엑셀 파일이 열려있습니다."
        except Exception as e:
            return False, f"저장 실패: {e}"

    def save_attachment(self, source_path, company_name, file_prefix):
        if not os.path.exists(source_path): return None, "파일 없음"
        try:
            current_year = datetime.now().strftime("%Y")
            year_dir = os.path.join(self.attachment_root, current_year)
            if not os.path.exists(year_dir): os.makedirs(year_dir)

            safe_name = "".join([c for c in str(company_name) if c.isalnum() or c in (' ', '_', '-')]).strip()
            comp_dir = os.path.join(year_dir, safe_name)
            if not os.path.exists(comp_dir): os.makedirs(comp_dir)

            fname = os.path.basename(source_path)
            name, ext = os.path.splitext(fname)
            today = datetime.now().strftime("%Y%m%d")
            new_name = f"{file_prefix}_{today}_{name}{ext}"
            
            dest_path = os.path.join(comp_dir, new_name)
            shutil.copy2(source_path, dest_path)
            return dest_path, None
        except Exception as e:
            return None, str(e)

    def add_log(self, action, details):
        try: user = getpass.getuser()
        except: user = "Unknown"
        new = {"일시": datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "작업자": user, "구분": action, "상세내용": details}
        new_df = pd.DataFrame([new])
        if self.df_log.empty: self.df_log = new_df
        else: self.df_log = pd.concat([self.df_log, new_df], ignore_index=True)
    
    def get_status_by_req_no(self, req_no):
        if self.df_data.empty: return None
        rows = self.df_data[self.df_data["관리번호"] == req_no]
        return rows.iloc[0]["Status"] if not rows.empty else None

    def get_filtered_data(self, status_list=None, keyword=""):
        df = self.df_data
        if df.empty: return df
        if status_list: df = df[df["Status"].isin(status_list)]
        if keyword:
            mask = pd.Series(False, index=df.index)
            for col in Config.SEARCH_TARGET_COLS:
                if col in df.columns:
                    mask |= df[col].astype(str).str.contains(keyword, case=False)
            df = df[mask]
        return df

    def set_dev_mode(self, enabled): self.is_dev_mode = enabled
    
    def create_backup(self):
        if not os.path.exists(self.current_excel_path): return False, "파일 없음"
        try:
            folder = os.path.dirname(self.current_excel_path)
            backup_folder = os.path.join(folder, "backup")
            if not os.path.exists(backup_folder): os.makedirs(backup_folder)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            fname = os.path.basename(self.current_excel_path)
            shutil.copy2(self.current_excel_path, os.path.join(backup_folder, f"{fname}_{timestamp}.bak"))
            return True, "백업 완료"
        except Exception as e: return False, str(e)

    def clean_old_logs(self): return True, "로그 정리 완료"

    def export_to_production_request(self, rows_data):
        prod_path = self.production_request_path
        if not os.path.exists(prod_path):
            return False, f"생산 요청 파일을 찾을 수 없습니다.\n경로: {prod_path}"

        try:
            wb = openpyxl.load_workbook(prod_path)
            if "Data" not in wb.sheetnames:
                return False, "'Data' 시트가 존재하지 않습니다."
            ws = wb["Data"]

            added_count = 0
            updated_count = 0

            for row_data in rows_data:
                client_name = row_data.get("업체명", "")
                client_note = "-"
                if not self.df_clients.empty:
                    c_row = self.df_clients[self.df_clients["업체명"] == client_name]
                    if not c_row.empty:
                        val = c_row.iloc[0].get("특이사항", "-")
                        if str(val) != "nan" and val: client_note = str(val)

                mgmt_no = str(row_data.get("관리번호", ""))
                model_name = str(row_data.get("모델명", ""))
                desc = str(row_data.get("Description", ""))

                mapping_values = [
                    mgmt_no, client_name, model_name, desc,
                    row_data.get("수량", 0), row_data.get("주문요청사항", ""), client_note,
                    row_data.get("수주일", ""), "-", "-", "-", "-", "-", "생산 접수",
                    row_data.get("발주서경로", ""), "-"
                ]

                target_row_idx = None
                for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    curr_mgmt = str(row[0]) if row[0] else ""
                    curr_model = str(row[2]) if row[2] else ""
                    curr_desc = str(row[3]) if row[3] else ""
                    
                    if curr_mgmt == mgmt_no and curr_model == model_name and curr_desc == desc:
                        target_row_idx = i
                        break
                
                if target_row_idx:
                    for col_idx, val in enumerate(mapping_values, start=1):
                        ws.cell(row=target_row_idx, column=col_idx, value=val)
                    updated_count += 1
                else:
                    ws.append(mapping_values)
                    added_count += 1

            wb.save(prod_path)
            wb.close()
            return True, f"신규: {added_count}건, 업데이트: {updated_count}건"

        except PermissionError:
            return False, "생산 요청 파일이 열려있습니다."
        except Exception as e:
            return False, f"생산 요청 내보내기 실패: {e}"